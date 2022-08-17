import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from fmtr import bars


wb = load_workbook("data/template.xlsx")
ws = wb.active

df = pd.read_csv(
    "data/crime.csv",
    encoding="utf-8",
    dtype={
        "INCIDENT_NUMBER": str,
        "OFFENSE_CODE": str,
        "OFFENSE_CODE_GROUP": str,
        "OFFENSE_DESCRIPTION": str,
        "DISTRICT": str,
        "REPORTING_AREA": str,
        "SHOOTING": str,
        "YEAR": str,
        "MONTH": str,
        "DAY_OF_WEEK": str,
        "HOUR": str,
    },
)

df1 = df[df["OFFENSE_CODE_GROUP"] == "Counterfeiting"]
df1 = df1.replace(np.nan, "N/A", regex=True)

total_crimes = len(df.index)
counterfeit = len(df1.index)
crimes_percentage = round(((counterfeit / total_crimes) * 100), 2)

ws["O8"].value = total_crimes
ws["P8"].value = counterfeit
ws["Q8"].value = crimes_percentage

bars()
df1["Count"] = 1
df2 = df1.groupby(["DISTRICT", "YEAR"]).count()["Count"]
print(df2)

bars()
df2 = df1.groupby(["DISTRICT", "YEAR"]).count()["Count"].unstack(level=0)
print(df2)

bars()
df2.drop(columns="N/A", inplace=True)
print(df2)

rows = dataframe_to_rows(df2)
for r_idx, row in enumerate(rows, 8):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save("data/crime_report.xlsx")
