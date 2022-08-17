from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    colors,
    Alignment,
    PatternFill,
    GradientFill,
    Border,
    Side,
    NamedStyle,
)

wb = Workbook()
ws = wb.active

for i in range(1, 20):
    ws.append(range(300))

ws.merge_cells("A1:B5")
ws.unmerge_cells("A1:B5")
ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)

# The lesson wrote: "colors.RED" but if you look down in the module, that usage is deprecated.
# So, COLOR_INDEX seems to align to http://dmcritchie.mvps.org/excel/colors.htm 🤷‍♀️ Ask excel 🤷‍♀️
cell = ws["B2"]
cell.font = Font(color=colors.COLOR_INDEX[43], size=20, italic=True)
cell.value = "Mike West did This"
cell.alignment = Alignment(horizontal="right", vertical="bottom")
cell.fill = GradientFill(stop=(colors.COLOR_INDEX[0], colors.COLOR_INDEX[25]))
wb.save("data/test_format.xlsx")

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True)
bd = Side(style="thick", color=colors.COLOR_INDEX[0])
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill("solid", fgColor="FFFF00")

count = 0
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
    col[count].style = highlight
    count = count + 1
wb.save("data/test_format_style.xlsx")
