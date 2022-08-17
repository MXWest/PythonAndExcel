from fmtr import bars
from openpyxl import load_workbook


wb = load_workbook("data/regions.xlsx")
ws = wb.active

bars()
cell_range = ws["A1":"C1"]  # Index by slicing row
print(cell_range)

bars()
col_c = ws["C"]
print(col_c)

bars()
col_range = ws["A":"C"]
print(col_range)

bars()
row_range = ws[1:5]
print(row_range)

bars()
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

bars()
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in row:
        print(cell)
