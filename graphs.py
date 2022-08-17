from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference


# Add chart to B14
wb = load_workbook("data/crime_report.xlsx")
ws = wb.active

# initial bar chart
barchart = BarChart()
data = Reference(ws, min_row=8, min_col=1, max_col=13, max_row=13)
barchart.add_data(data, titles_from_data=True)
ws.add_chart(barchart, "B14")
wb.save("data/crime_report_barchart.xlsx")

# clean up the chart a bit
wb2 = load_workbook("data/crime_report.xlsx")
ws2 = wb2.active

data = Reference(ws, min_row=10, min_col=1, max_col=13, max_row=13)
labels = Reference(ws, min_row=8, min_col=2, max_row=8, max_col=13)
barchart.add_data(data, from_rows=True, titles_from_data=True)
barchart.set_categories(labels)
barchart.title = "Counterfeit Crimes by District"
barchart.height = (
    4.56  # Instructor explains the calculations used to get height and width
)
barchart.width = 20.3

piechart = PieChart()
data = Reference(ws, min_col=15, max_col=16, min_row=8, max_row=8)
labels = Reference(ws, min_col=15, max_col=16, min_row=7, max_row=7)
piechart.add_data(data, from_rows=True)
piechart.set_categories(labels)
piechart.height = barchart.height
piechart.width = 8.45

ws2.add_chart(barchart, "B14")
ws2.add_chart(piechart, "N14")
wb2.save("data/crime_report_barchart2.xlsx")
