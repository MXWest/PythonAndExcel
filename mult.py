import pandas as pd
from fmtr import bars
from openpyxl import load_workbook
from openpyxl.styles import Font


df1 = pd.read_excel("data/shifts.xlsx", sheet_name="Sheet")
df2 = pd.read_excel("data/shifts.xlsx", sheet_name="Sheet1")
df3 = pd.read_excel("data/shift_3.xlsx")

bars()
df_all = pd.concat([df1, df2, df3], sort=False)
print(df_all)

bars()
print(df_all.loc[50])

bars()
print(df_all.groupby(["Shift"]).mean()["Units Sold"])

bars()
df_all.to_excel("data/all_shifts.xlsx", index=None)

bars()
wb = load_workbook("data/all_shifts.xlsx")
ws = wb.active

total_col = ws["G1"]
total_col.font = Font(bold=True)
total_col.value = "Total"

e_col, f_col = ["E", "F"]
for row in range(2, 300):
    result_cell = "G{}".format(row)
    e_value = ws[e_col + str(row)].value
    f_value = ws[f_col + str(row)].value
    ws[result_cell] = e_value + f_value

wb.save("data/mike_totaled.xlsx")
