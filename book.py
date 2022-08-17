from openpyxl import Workbook
from openpyxl import load_workbook


def bars() -> None:
    print("-----")


wb = Workbook()  # initialize new Workbook to work with
ws = wb.active  # worksheet

bars()
ws1 = wb.create_sheet("Sheet One by Mike")
ws2 = wb.create_sheet("Mike Sheet Zero", 0)
ws.title = "The First One I Created"
print(wb.sheetnames)

bars()
wb2 = load_workbook("data/regions.xlsx")
new_sheet = wb2.create_sheet("Fourth Sheet by Mike")
active_sheet = wb2.active
cell = active_sheet["A4"]
print(cell)
print(cell.value)
active_sheet["A4"] = "us-east1"
wb2.save("data/regions-mod1.xlsx")
