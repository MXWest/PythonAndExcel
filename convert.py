import pandas as pd
from fmtr import bars
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook("data/regions.xlsx")
ws = wb.active
df = pd.read_excel("data/all_shifts.xlsx")

bars()
df1 = df[["Sales Rep", "Cost per", "Units Sold"]]
df1["Total"] = df1["Cost per"] * df1["Units Sold"]  # Creates a new column, "Total"
print(df1)

bars()
rows = dataframe_to_rows(df1, index=False)
print(rows)

bars()
for row in rows:
    print(row)

bars()
for row in rows:
    for col in row:
        print(col)

bars()
rows = dataframe_to_rows(df1, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, col in enumerate(row, 6):
        print(col)
bars()
rows = dataframe_to_rows(df1, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, col in enumerate(row, 6):
        ws.cell(row=r_idx, column=c_idx, value=col)

wb.save("data/combined_mike.xlsx")
